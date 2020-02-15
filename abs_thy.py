from selenium import webdriver
import time
import openpyxl
import os

#web sayfasi acilacak
#web sayfasindaki belirli noktaya awb nosu girilecek
#list dugmesine basilacak

os.chdir('/users/akin/desktop')
#Buraya kadar dosyaya isim vermedik ama olusturabildik...
#Cunku suanda sadece Python icinde olustu daha kayit etmedik
workbook = openpyxl.Workbook()
workbook.get_sheet_names()
sheet = workbook.get_sheet_by_name('Sheet')



def thy_website_check(awb_number):

    adress = 'https://www.turkishcargo.com.tr/en/online-services/shipment-tracking'
    browser = webdriver.Chrome()
    browser.get(adress)

    awb_searchfield = browser.find_element_by_css_selector('#serviceform > div:nth-child(6) > div.col-md-4.col-sm-4.col-xs-8 > div > div > input')
    time.sleep(3)
    awb_filled_searchfield = awb_searchfield.send_keys(awb_number)

    list_button = browser.find_element_by_css_selector('#serviceform > div:nth-child(8) > div:nth-child(2) > button')
    list_button.click()
    time.sleep(3)

    last_movement = browser.find_element_by_css_selector('#accordion > li > div.submenu > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(2)')
    last_movement = last_movement.text

    print(last_movement)

    
    
    for i in range(1, 3):
        os.chdir('/users/akin/desktop')
        workbook_results = openpyxl.load_workbook('thy_tracking.xlsx')
        sheet1 = workbook_results.get_sheet_by_name('Sheet')

        last_movement = sheet1.cell(row=i, column=1).value
        
        workbook_results.save('thy_tracking.xlsx')


os.chdir('/users/akin/desktop')
awb_workbook = openpyxl.load_workbook('awb_numbers.xlsx')
sheet1 = awb_workbook.get_sheet_by_name('Sheet1')

for i in range(1, 3):
    cell = sheet1.cell(row=i, column=1).value
    thy_website_check(cell)


#sheet['A1'].value = last_movement

#workbook.save('thy_tracking.xlsx')
